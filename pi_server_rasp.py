#!/usr/bin/env python
# -*- coding: cp1252 -*-
import socket
import time
try:
    # Python2
    import Tkinter as tk
except ImportError:
    # Python3
    import tkinter as tk
import os
import sys
import openpyxl
import subprocess
import os.path
import threading
from socket import AF_INET, SOCK_STREAM
tk_instance_exist=False
root=0
text=0
myproc=0
HOST = '' 
PORT = 8080
s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
s.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
s.bind((HOST, PORT))
s.listen(1)
#while 1:
#    s.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
#    s.bind((HOST, PORT))
#    s.listen(8080)
#    conn,addr = s.accept()
#    data=conn.recv(1024)
#    print(data)
#    
#    conn.sendall()
#    conn.close()


def handle(conn,data):
    global myproc
    global tk_instance_exist
    global root
    global text
    #reply=raw_input();
    parseData=data.split("|")
    if(parseData[0]=="video1"):
        reply="success"
        conn.sendall(reply.encode())
        conn.close()
	if not (myproc==0):
		myproc.kill()
		myproc.stdin.write('q')
	myproc=subprocess.Popen(['./video1.sh'], stdin=subprocess.PIPE)
        #myproc=subprocess.Popen(["omxplayer --win '0 0 300 300'","/home/pi/video/ad1.mp4"],stdin=subprocess.PIPE)
    elif(parseData[0]=="video2"):
        reply="success"
        conn.sendall(reply.encode())
        conn.close()
	if not (myproc==0):
		myproc.kill()
		myproc.stdin.write('q')
	myproc=subprocess.Popen(['./video2.sh'], stdin=subprocess.PIPE)
        #myproc=subprocess.Popen(["omxplayer --win '0 0 300 300'","/home/pi/video/ad3.mp4"],stdin=subprocess.PIPE)
    elif(parseData[0]=="video3"):
        reply="success"
        conn.sendall(reply.encode())
        conn.close()
	if not (myproc==0):
		myproc.kill()
		myproc.stdin.write('q')
	myproc=subprocess.Popen(['./video3.sh'],stdin=subprocess.PIPE)
        #myproc=subprocess.Popen(["omxplayer --win '0 0 300 300'","/home/pi/video/ad5.mp4"],stdin=subprocess.PIPE)
    elif(parseData[0]=="videoAll"):
        reply="success"
        conn.sendall(reply.encode())
        conn.close()
	if not (myproc==0):
		myproc.kill()
		myproc.stdin.write('q')
	myproc=subprocess.Popen(['./video.sh'],stdin=subprocess.PIPE)
        #myproc=subprocess.Popen(["omxplayer --win '0 0 300 300'","/home/pi/ad3.mp4"],stdin=subprocess.PIPE)
    elif(parseData[0]=="public bloodrecords"):
        if not (os.path.isfile('blood_donation.xlsx')==False):
            wb = openpyxl.load_workbook('blood_donation.xlsx')
            worksheet = wb.worksheets[0]
            cell_value=""
            reply=""
            if not (worksheet.max_row==1):
                reply="{\"Data\":["
                #here you iterate over the rows in the specific column
                for row in range(2,worksheet.max_row+1):
                    reply=reply+"{"
                    for column in "ABCDEF":  #Here you can add or reduce the columns
                        cell_name = "{}{}".format(column, row)
                        cell_value=worksheet[cell_name].value # the value of the specific cell
                        if(column=="A"):
                            reply=reply+"\"name\":\""+cell_value+"\","
                        elif(column=="B"):
                            reply=reply+"\"number\":\""+cell_value+"\","
                        elif(column=="C"):
                            reply=reply+"\"cnic\":\""+cell_value+"\","
                        elif(column=="D"):
                            reply=reply+"\"bottles\":\""+cell_value+"\","
                        elif(column=="E"):
                            reply=reply+"\"bloodgroup\":\""+cell_value+"\","
                        elif(column=="F"):
                            reply=reply+"\"date\":\""+cell_value+"\""
                    reply=reply+"}"
                    if not(row==(worksheet.max_row)):
                        reply=reply+","
                    else:
                        reply=reply+"]}"
            else:
                reply="null"
            conn.sendall(reply.encode())
            conn.close()
            #p = subprocess.Popen(["C:/Program Files/Windows Media Player/wmplayer.exe","file:///E:/Android Backup/Digital signage/ad5.mp4"])
        else:
            reply="null"
            conn.sendall(reply.encode())
            conn.close()
    elif(parseData[0]=="public cashrecords"):
        if not (os.path.isfile('cash_donation.xlsx')==False):
            wb = openpyxl.load_workbook('cash_donation.xlsx')
            worksheet = wb.worksheets[0]
            cell_value=""
            reply=""
            if not (worksheet.max_row==1):
                reply="{\"Data\":["
                #here you iterate over the rows in the specific column
                for row in range(2,worksheet.max_row+1):
                    reply=reply+"{"
                    for column in "ABCD":  #Here you can add or reduce the columns
                        cell_name = "{}{}".format(column, row)
                        cell_value=worksheet[cell_name].value # the value of the specific cell
                        if(column=="A"):
                            reply=reply+"\"name\":\""+cell_value+"\","
                        elif(column=="B"):
                            reply=reply+"\"number\":\""+cell_value+"\","
                        elif(column=="C"):
                            reply=reply+"\"cnic\":\""+cell_value+"\","
                        elif(column=="D"):
                            reply=reply+"\"amount\":\""+cell_value+"\""
                    reply=reply+"}"
                    if not(row==(worksheet.max_row)):
                        reply=reply+","
                    else:
                        reply=reply+"]}"
            else:
                reply="null"
            conn.sendall(reply.encode())
            conn.close()
            #p = subprocess.Popen(["C:/Program Files/Windows Media Player/wmplayer.exe","file:///E:/Android Backup/Digital signage/ad5.mp4"])
        else:
            reply="null"
            conn.sendall(reply.encode())
            conn.close()
    elif(parseData[0]=="personal cashrecords"):
        if not (os.path.isfile('cash_donation.xlsx')==False):
            Id=parseData[1]
            wb = openpyxl.load_workbook('cash_donation.xlsx')
            worksheet = wb.worksheets[0]
            cell_value=""
            reply=""
            if not (worksheet.max_row==1):
                reply="{\"Data\":["
                #here you iterate over the rows in the specific column
                for row in range(2,worksheet.max_row+1):
                    if(worksheet["{}{}".format("E", row)].value==Id):
                        reply=reply+"{"
                        for column in "ABCD":  #Here you can add or reduce the columns
                            cell_name = "{}{}".format(column, row)
                            cell_value=worksheet[cell_name].value # the value of the specific cell
                            if(column=="A"):
                                reply=reply+"\"name\":\""+cell_value+"\","
                            elif(column=="B"):
                                reply=reply+"\"number\":\""+cell_value+"\","
                            elif(column=="C"):
                                reply=reply+"\"cnic\":\""+cell_value+"\","
                            elif(column=="D"):
                                reply=reply+"\"amount\":\""+cell_value+"\""
                        reply=reply+"}"
                        if not(row==(worksheet.max_row)):
                            reply=reply+","
                        else:
                            reply=reply+"]}"
                if(reply=="{\"Data\":["):
                   reply="null" 
            else:   
                reply="null"
            conn.sendall(reply.encode())
            conn.close()
        else:
            reply="null"
            conn.sendall(reply.encode())
            conn.close()
    elif(parseData[0]=="personal bloodrecords"):
        if not (os.path.isfile('blood_donation.xlsx')==False):
            Id=parseData[1]
            wb = openpyxl.load_workbook('blood_donation.xlsx')
            worksheet = wb.worksheets[0]
            cell_value=""
            reply=""
            if not (worksheet.max_row==1):
                reply="{\"Data\":["
                #here you iterate over the rows in the specific column
                for row in range(2,worksheet.max_row+1):
                    if(worksheet["{}{}".format("G", row)].value==Id):
                        reply=reply+"{"
                        for column in "ABCDEF":  #Here you can add or reduce the columns
                            cell_name = "{}{}".format(column, row)
                            cell_value=worksheet[cell_name].value # the value of the specific cell
                            if(column=="A"):
                                reply=reply+"\"name\":\""+cell_value+"\","
                            elif(column=="B"):
                                reply=reply+"\"number\":\""+cell_value+"\","
                            elif(column=="C"):
                                reply=reply+"\"cnic\":\""+cell_value+"\","
                            elif(column=="D"):
                                reply=reply+"\"bottles\":\""+cell_value+"\","
                            elif(column=="E"):
                                reply=reply+"\"bloodgroup\":\""+cell_value+"\","
                            elif(column=="F"):
                                reply=reply+"\"date\":\""+cell_value+"\""
                        reply=reply+"}"
                        if not(row==(worksheet.max_row)):
                            reply=reply+","
                        else:
                            reply=reply+"]}"
                if(reply=="{\"Data\":["):
                    reply="null" 
            else:
                reply="null"
            conn.sendall(reply.encode())
            conn.close()
        else:
            reply="null"
            conn.sendall(reply.encode())
            conn.close()
    elif(parseData[0]=="cash donation"):
        try:
            if(os.path.isfile('cash_donation.xlsx')==False):
                wb = openpyxl.Workbook()
                sheet = wb.worksheets[0]
                sheet.column_dimensions['A'].width = 30
                sheet.column_dimensions['B'].width = 30
                sheet.column_dimensions['C'].width = 30
                sheet.column_dimensions['D'].width = 30
                sheet.column_dimensions['E'].width = 20
                sheet.cell(1, 1, 'Name')
                sheet.cell(row=1, column=1).font = sheet.cell(row=1, column=1).font.copy(bold=True)
                sheet.cell(1, 2, 'Phone Number')
                sheet.cell(row=1, column=2).font = sheet.cell(row=1, column=1).font.copy(bold=True)
                sheet.cell(1, 3, 'Account Number')
                sheet.cell(row=1, column=3).font = sheet.cell(row=1, column=1).font.copy(bold=True)
                sheet.cell(1, 4, 'Cash Donation')
                sheet.cell(row=1, column=4).font = sheet.cell(row=1, column=1).font.copy(bold=True)
                sheet.cell(1, 5, 'ID')
                sheet.cell(row=1, column=5).font = sheet.cell(row=1, column=1).font.copy(bold=True)
                wb.save('cash_donation.xlsx')
            wb = openpyxl.load_workbook('cash_donation.xlsx')
            sheet = wb.worksheets[0]
            row = sheet.max_row + 1
            column_count = sheet.max_column
            name=parseData[1]
            number=parseData[2]
            cnic=parseData[3]
            amount=parseData[4]
            Id=parseData[7]
            new_row = [name, number,cnic, amount,Id]

            for col, entry in enumerate(new_row, start=1):
                sheet.cell(row=row, column=col, value=entry)

            wb.save('cash_donation.xlsx')
            reply="success"
            conn.sendall(reply.encode())
            conn.close()
            tickerText="Name : " + name + ",    Phone Number : " +number+ ",    Donation : Rs." + amount + "."
            
            if(tk_instance_exist==False):
                root = tk.Tk()
                root.title('Donation Information')
                #root = tk.Tk()
                w = root.winfo_screenwidth() # width for the Tk root
                h = 66 # height for the Tk root

                # get screen width and height
                ws = root.winfo_screenwidth() # width of the screen
                hs = root.winfo_screenheight() # height of the screen

                # calculate x and y coordinates for the Tk root window
                x = (ws/1) - (w/1)
                y = (hs/1.1) - (h/1)

                # set the dimensions of the screen 
                # and where it is placed
                root.geometry('%dx%d+%d+%d' % (w, h, x, y))
                root.attributes('-topmost',True)
                # width --> width in chars, height --> lines of text
                text_width = 100
                text = tk.Text(root, width=text_width, height=1, bg='blue')
                text.pack()
                # use a proportional font to handle spaces correctly
                text.config(font=('aerial', 40, 'bold'))
                tk_instance_exist=True
                print('not created')
            else:
                print('already')
	    text_width = 100
 	    newsText = tickerText
 	    # pad front and end of text with spaces
	    paddingText = ' ' * text_width
    	    # concatenate it all
	    s = paddingText +  newsText  + paddingText
	    for k in range(len(s)):
	        # use string slicing to do the trick
	        ticker_text = s[k:k+text_width]
	        text.insert("1.0", ticker_text)
	        root.update()
 	       # delay by 0.22 seconds
	        time.sleep(0.12)
	        if(k==(len(s)-70)):
 	           #root.iconify()
 	           break
 	    root.mainloop() 
            #threading.Thread(target=displayTickerWindow, args=(tickerText,root,text,)).start()
            #displayTickerWindow(tickerText)
        except:
            reply="Unable to add record, plz try again."
            conn.sendall(reply.encode())
            conn.close()
        # Create an new Excel file and add a worksheet.
        #workbook = xlsxwriter.Workbook('demo.xlsx')
        #worksheet = workbook.add_worksheet()

        # Widen the first column to make the text clearer.
        #worksheet.set_column('A:A', 20)
        #worksheet.set_column('B:B', 20)
        #worksheet.set_column('C:C', 20)

        # Add a bold format to use to highlight cells.
        #bold = workbook.add_format({'bold': True})

        # Write some simple text.
        #worksheet.write('A1', 'Name', bold)
        #worksheet.write('B1', 'Phone Number', bold)
        #worksheet.write('C1', 'Donation Amount', bold)

        # Write some numbers, with row/column notation.
        #worksheet.write(row_count, 0, 123)
        #worksheet.write(row_count, 1, 123.456)
        #worksheet.write(row_count, 2, 123.456)

        #workbook.close()
        
        
    elif(parseData[0]=="blood donation"):
        try:
            if(os.path.isfile('blood_donation.xlsx')==False):
                wb = openpyxl.Workbook()
                sheet = wb.worksheets[0]
                sheet.column_dimensions['A'].width = 30
                sheet.column_dimensions['B'].width = 30
                sheet.column_dimensions['C'].width = 30
                sheet.column_dimensions['D'].width = 30
                sheet.column_dimensions['E'].width = 30
                sheet.column_dimensions['F'].width = 30
                sheet.column_dimensions['G'].width = 20
                sheet.cell(1, 1, 'Name')
                sheet.cell(row=1, column=1).font = sheet.cell(row=1, column=1).font.copy(bold=True)
                sheet.cell(1, 2, 'Phone Number')
                sheet.cell(row=1, column=2).font = sheet.cell(row=1, column=1).font.copy(bold=True)
                sheet.cell(1, 3, 'CNIC')
                sheet.cell(row=1, column=3).font = sheet.cell(row=1, column=1).font.copy(bold=True)
                sheet.cell(1, 4, 'Blood Bottle Donation')
                sheet.cell(row=1, column=4).font = sheet.cell(row=1, column=1).font.copy(bold=True)
                sheet.cell(1, 5, 'Blood Group')
                sheet.cell(row=1, column=5).font = sheet.cell(row=1, column=1).font.copy(bold=True)
                sheet.cell(1, 6, 'Last Donated Blood Date')
                sheet.cell(row=1, column=6).font = sheet.cell(row=1, column=1).font.copy(bold=True)
                sheet.cell(1, 7, 'ID')
                sheet.cell(row=1, column=7).font = sheet.cell(row=1, column=1).font.copy(bold=True)
                wb.save('blood_donation.xlsx')
            wb = openpyxl.load_workbook('blood_donation.xlsx')
            sheet = wb.worksheets[0]
            row = sheet.max_row + 1
            column_count = sheet.max_column
            name=parseData[1]
            number=parseData[2]
            cnic=parseData[3]
            amount=parseData[4]
            bloodgroup=parseData[5]
            lastdate=parseData[6]
            Id=parseData[7]
            new_row = [name, number, cnic, amount, bloodgroup, lastdate,Id ]

            for col, entry in enumerate(new_row, start=1):
                sheet.cell(row=row, column=col, value=entry)

            wb.save('blood_donation.xlsx')
            reply="success"
            conn.sendall(reply.encode())
            conn.close()
            tickerText="Name : " + name + ",    Phone Number : " +number+ ",    Donation : " + amount + " bottles of blood."
            
            if(tk_instance_exist==False):
                root = tk.Tk()
                root.title('Donation Information')
                #root = tk.Tk()
                w = root.winfo_screenwidth() # width for the Tk root
                h = 66 # height for the Tk root

                # get screen width and height
                ws = root.winfo_screenwidth() # width of the screen
                hs = root.winfo_screenheight() # height of the screen

                # calculate x and y coordinates for the Tk root window
                x = (ws/1) - (w/1)
                y = (hs/1.1) - (h/1)

                # set the dimensions of the screen 
                # and where it is placed
                root.geometry('%dx%d+%d+%d' % (w, h, x, y))
                root.attributes('-topmost',True)
                # width --> width in chars, height --> lines of text
                text_width = 100
                text = tk.Text(root, width=text_width, height=1, bg='blue')
                text.pack()
                # use a proportional font to handle spaces correctly
                text.config(font=('aerial', 40, 'bold'))
                tk_instance_exist=True
                print('not created')
            else:
                print('already')   
    	    text_width = 100
 	    newsText = tickerText
 	    # pad front and end of text with spaces
	    paddingText = ' ' * text_width
    	    # concatenate it all
	    s = paddingText +  newsText  + paddingText
	    for k in range(len(s)):
	        # use string slicing to do the trick
	        ticker_text = s[k:k+text_width]
	        text.insert("1.0", ticker_text)
	        root.update()
 	       # delay by 0.22 seconds
	        time.sleep(0.12)
	        if(k==(len(s)-70)):
 	           #root.iconify()
 	           break
 	    root.mainloop() 
            #threading.Thread(target=displayTickerWindow, args=(tickerText,root,text,)).start()
            #displayTickerWindow(tickerText)
        except:
            reply="Unable to add record, plz try again."
            conn.sendall(reply.encode())
            conn.close()
        
    else:
        reply="unable to communicate with server"
        conn.sendall(reply.encode())
        conn.close()


def displayTickerWindow(tickerText,root,text):
    print('start ticker')
    
    text_width = 100
    newsText = tickerText
    # pad front and end of text with spaces
    paddingText = ' ' * text_width
    # concatenate it all
    s = paddingText +  newsText  + paddingText
    for k in range(len(s)):
        # use string slicing to do the trick
        ticker_text = s[k:k+text_width]
        text.insert("1.0", ticker_text)
        root.update()
        # delay by 0.22 seconds
        time.sleep(0.12)
        if(k==(len(s)-90)):
            #root.iconify()
            break
    root.mainloop()    

while True:
    print('listening to client...')
    conn, addr = s.accept()
    data=conn.recv(1024)
    
    print(data)
    #print('handling connection from %s' % (addr,))
    threading.Thread(target=handle, args=(conn,data,)).start()
